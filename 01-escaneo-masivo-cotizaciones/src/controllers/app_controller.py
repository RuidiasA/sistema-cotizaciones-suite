"""Controlador Principal de la Aplicación de Cotizaciones.

Actúa como orquestador central (Patrón MVC) vinculando las interacciones
de la vista (MainView) con los servicios de dominio (ExcelScanService, 
BenchmarkingService, QuoteService, VariationService) y actualizando el estado de la UI.

Principios aplicados:
    - Single Responsibility (SRP): Exclusivamente dedicado a coordinar el flujo de datos y tareas asíncronas.
    - Dependency Inversion (DIP): Se apoya en abstraer la lógica de cálculo en servicios especializados.
"""

from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
import os
import threading
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models.constants import DEFAULT_EXCEL_FOLDER, HOJAS_EXCLUIDAS, MACRO_CATEGORIAS
from ..models.entities import BenchmarkingMatrix, FileScanReport, PriceStats, ScanRow
from ..services.benchmarking_service import BenchmarkingService
from ..services.excel_scan_service import ExcelScanService
from ..services.quote_service import QuoteService
from ..services.variation_service import VariationService
from ..views.main_view import MainView


class AppController:
    """Controlador orquestador de la suite de cotizaciones y benchmarking."""

    def __init__(self) -> None:
        """Inicializa los servicios de dominio, el estado interno y la vista raíz."""
        self._variation_service = VariationService(MACRO_CATEGORIAS)
        self._scan_service = ExcelScanService(HOJAS_EXCLUIDAS)
        self._quote_service = QuoteService()
        self._benchmarking_service = BenchmarkingService()

        self._stats = PriceStats()
        self._matched_total = 0
        self._last_scan_rows: List[ScanRow] = []
        self._current_benchmarking: Optional[BenchmarkingMatrix] = None
        self._last_keyword: str = ""
        self._last_scan_folder: str = ""
        self._enable_raw_debug_export = True

        self._stop_event = threading.Event()
        self._executor = ThreadPoolExecutor(max_workers=2)

        self._view = MainView(
            controller=self,
            categories=self._variation_service.get_categories(),
            default_folder=DEFAULT_EXCEL_FOLDER,
            on_quote=self.handle_quote,
            on_scan=self.handle_scan,
            on_cancel=self.handle_cancel,
        )

    def run(self) -> None:
        """Inicia el bucle principal de eventos de la interfaz gráfica."""
        self._view.mainloop()

    def has_scan_data(self) -> bool:
        """Indica si existen datos de un escaneo previo en memoria.

        Returns:
            True si hay filas procesadas disponibles, False en caso contrario.
        """
        return bool(self._last_scan_rows)

    def handle_scan(self, folder: str, categoria: str, keyword: str) -> None:
        """Inicia la tarea asíncrona de escaneo de archivos Excel."""
        self._stop_event.clear()
        self._last_keyword = keyword.strip()
        self._last_scan_folder = folder

        self._view.set_scanning_state(True)
        self._view.enable_export(False)
        self._view.set_status("Iniciando escaneo...")
        self._view.clear_results()

        search_pack = self._build_search_pack(categoria, keyword)

        def _on_progress(completed: int, total: int, file_name: str, eta_str: str) -> None:
            # Actualizar la UI cada 5 archivos o al finalizar para no saturar el mainloop
            if completed % 5 == 0 or completed == total:
                def update_ui_progress() -> None:
                    self._view.set_status(f"Procesando ({completed}/{total}) - ETA {eta_str}")
                    self._view.append_log(f"📄 [{completed}/{total}] {file_name} (ETA: {eta_str})")
                
                self._view.after(0, update_ui_progress)

        future = self._executor.submit(
            self._scan_service.scan_folder, 
            folder, 
            search_pack, 
            self._stop_event,
            _on_progress
        )
        future.add_done_callback(lambda f: self._on_scan_done(f))

    def _build_search_pack(self, categoria: str, keyword: str) -> dict:
        categoria = (categoria or "").strip()
        keyword = (keyword or "").strip()

        if not categoria:
            search_pack = self._variation_service.get_global_search_pack()
            return self._dedupe_search_pack(search_pack, keyword)

        macro, sub = self._resolve_macro_subcategoria(categoria)
        if not macro:
            search_pack = self._variation_service.get_global_search_pack()
            return self._dedupe_search_pack(search_pack, keyword)

        tags: List[str] = []
        excludes: List[str] = []
        macro_info = MACRO_CATEGORIAS.get(macro, {})
        global_excludes = macro_info.get("global_exclude", [])
        subcategorias = macro_info.get("subcategorias", {})

        if sub:
            clusters = subcategorias.get(sub, [])
        else:
            clusters = []
            for sub_clusters in subcategorias.values():
                clusters.extend(sub_clusters)

        for cluster in clusters:
            tags.extend(cluster.get("tags", []))
            excludes.extend(cluster.get("exclude", []))

        excludes.extend(global_excludes)
        return self._dedupe_search_pack({"tags": tags, "exclude": excludes}, keyword)

    def _dedupe_search_pack(self, search_pack: dict, keyword: str) -> dict:
        tags = search_pack.get("tags", [])
        excludes = search_pack.get("exclude", [])

        tags_set = set(tags)
        excludes_set = set(excludes)
        if keyword:
            tags_set.add(keyword)

        return {
            "tags": list(tags_set),
            "exclude": list(excludes_set),
        }

    def _resolve_macro_subcategoria(self, categoria: str) -> tuple[str | None, str | None]:
        if not categoria:
            return None, None

        if ">" in categoria:
            parts = [part.strip() for part in categoria.split(">", 1)]
            if len(parts) == 2 and parts[0] and parts[1]:
                return parts[0], parts[1]

        if categoria in MACRO_CATEGORIAS:
            return categoria, None

        for macro, info in MACRO_CATEGORIAS.items():
            subcategorias = info.get("subcategorias", {})
            if categoria in subcategorias:
                return macro, categoria

        return None, None

    def handle_cancel(self) -> None:
        """Notifica al motor de escaneo la activación de la señal de parada."""
        self._stop_event.set()
        self._view.append_log("🛑 Cancelando escaneo... (esperando cierre de hilo)")
        self._view.set_status("Cancelando")

    def _on_scan_done(self, future: object) -> None:
        try:
            reports: List[FileScanReport] = future.result()
        except Exception as exc:
            self._view.after(
                0,
                lambda: (
                    self._view.set_scanning_state(False),
                    self._view.set_status(f"Error: {exc}"),
                ),
            )
            return

        def update_ui() -> None:
            self._stats = PriceStats()
            filas_globales: List[ScanRow] = []

            for report in reports:
                if report.error_message:
                    self._view.append_log(f"⚠️ {report.file_name}: {report.error_message}")
                    continue

                self._view.append_log(f"✅ {report.file_name} ({report.sheet_name})")
                filas_globales.extend(report.matched_rows)
                self._stats.merge(report.stats)

            total_files = len(reports)
            valid_files = sum(1 for r in reports if not r.error_message)
            invalid_files = total_files - valid_files

            sheets_valid = 0
            for r in reports:
                if not r.error_message and r.sheet_name:
                    hojas = [s.strip() for s in str(r.sheet_name).split(",") if s and s.strip()]
                    sheets_valid += len(hojas)

            total_rows_generated = sum(len(r.matched_rows) for r in reports)

            self._view.append_log("")
            self._view.append_log(
                f"Resumen: archivos detectados: {total_files} | válidos: {valid_files} | inválidos: {invalid_files}"
            )
            self._view.append_log(
                f"Hojas válidas leídas: {sheets_valid} | Filas generadas: {total_rows_generated}"
            )

            filas_globales.sort(key=lambda x: str(x.articulo).strip().lower())
            self._view.add_rows(filas_globales)
            self._last_scan_rows = list(filas_globales)

            if self._enable_raw_debug_export:
                raw_df = self._scan_service.get_last_raw_scan_dataframe()
                if not raw_df.empty:
                    def _async_debug_save():
                        try:
                            output_folder = self._last_scan_folder or os.getcwd()
                            output_file = self._scan_service.export_raw_scan_dataframe(raw_df, output_folder)
                            self._view.after(0, lambda: self._view.append_log(f"🧪 Debug raw export: {output_file}"))
                        except Exception as exc:
                            self._view.after(0, lambda: self._view.append_log(f"⚠️ Debug raw export falló: {exc}"))
                    
                    threading.Thread(target=_async_debug_save, daemon=True).start()

            self._matched_total = len(filas_globales)
            self._view.set_stats_text(self._format_stats(self._matched_total))

            if self._stop_event.is_set():
                self._view.set_status("Escaneo Cancelado")
            else:
                self._view.set_status("Listo")

            self._view.set_scanning_state(False)

        self._view.after(0, update_ui)

    def _format_stats(self, matched_total: int) -> str:
        if matched_total == 0:
            return (
                "Margen Promedio (>= 1000): 35.00% (0 casos) | "
                "Margen Promedio (>= 500): 35.00% (0 casos) | "
                "Margen Promedio (Resto): 35.00% (0 casos)"
            )
        avg_gt_1000 = self._stats.promedio_para_cantidad(1000)
        avg_gt_500 = self._stats.promedio_para_cantidad(500)
        avg_rest = self._stats.promedio_para_cantidad(1)
        return (
            "Margen Promedio (>= 1000): "
            f"{avg_gt_1000:.2f}% ({self._stats.count_gt_1000} casos) | "
            "Margen Promedio (>= 500): "
            f"{avg_gt_500:.2f}% ({self._stats.count_gt_500} casos) | "
            "Margen Promedio (Resto): "
            f"{avg_rest:.2f}% ({self._stats.count_rest} casos)"
        )

    def handle_quote(self, product_name: str, cantidad: int, precio_prov: float) -> None:
        """Procesa una solicitud de la calculadora rápida de cotizaciones.

        Args:
            product_name: Descripción o nombre del producto.
            cantidad: Volumen a cotizar.
            precio_prov: Costo otorgado por el proveedor.
        """
        arquetipo_buscado = self._benchmarking_service.extraer_arquetipo(product_name)
        margen_final = 35.0
        encontrado_en_benchmarking = False

        if self._current_benchmarking and arquetipo_buscado:
            arq_data = self._current_benchmarking.get_arquetipo_por_nombre(arquetipo_buscado)
            if arq_data:
                margen_final = self._current_benchmarking.get_margen_para_cantidad(
                    arquetipo_buscado, cantidad
                )
                encontrado_en_benchmarking = True

        if encontrado_en_benchmarking:
            precio_unit = round(precio_prov * (1 + (margen_final / 100.0)), 2)
            total = round(precio_unit * cantidad, 2)
            result = {"margen": margen_final, "precio_unit": precio_unit, "total": total}
        else:
            stats = self._stats if self._matched_total > 0 else PriceStats()
            result = self._quote_service.create_quote(
                product_name,
                cantidad,
                precio_prov,
                stats,
                margen_defecto=margen_final,
            )

        known = encontrado_en_benchmarking or self._variation_service.is_known_product(product_name)

        if encontrado_en_benchmarking:
            self._view.after(
                0,
                lambda: self._view.append_log(
                    f"💡 Calculando '{product_name}' con margen de benchmarking: {margen_final}%"
                ),
            )

        self._view.after(0, lambda: self._view.show_quote_result(result, known))
        self._view.after(
            0,
            lambda: self._view.update_quote_cards(
                result["margen"], result["precio_unit"], result["total"]
            ),
        )

    def handle_benchmarking(self, categoria: str) -> None:
        """Inicia el cálculo de la matriz de benchmarking para una categoría.

        Args:
            categoria: Nombre de la categoría o subcategoría objetivo.
        """
        if not self._last_scan_rows:
            self._view.after(0, lambda: self._view.set_status("Sin datos para benchmarking"))
            return

        rows_para_benchmarking = self._last_scan_rows
        categoria_limpia = (categoria or "").strip()
        if categoria_limpia:
            rows_para_benchmarking = [
                row
                for row in self._last_scan_rows
                if self._variation_service.matches_category(categoria_limpia, row.articulo)
            ]

        if not rows_para_benchmarking:
            self._view.after(0, lambda: self._view.clear_results())
            self._view.after(
                0, lambda: self._view.set_status(f"Sin datos para la categoría: {categoria_limpia}")
            )
            self._view.after(0, lambda: self._view.enable_export(False))
            return

        self._view.after(0, lambda: self._view.set_benchmarking_state(True))
        self._view.after(0, lambda: self._view.enable_export(False))
        self._view.after(0, lambda: self._view.set_status("Generando Benchmarking..."))

        future = self._executor.submit(
            self._benchmarking_service.generar_benchmarking,
            list(rows_para_benchmarking),
            categoria_limpia,
            self._last_keyword,
        )
        future.add_done_callback(lambda f: self._on_benchmarking_done(f))

    def _on_benchmarking_done(self, future: object) -> None:
        try:
            matrix: BenchmarkingMatrix = future.result()
        except Exception as exc:
            self._view.after(0, lambda: self._view.set_benchmarking_state(False))
            self._view.after(0, lambda: self._view.set_status(f"Error: {exc}"))
            return

        def update_ui() -> None:
            self._current_benchmarking = matrix
            benchmarking_rows = self._benchmarking_rows_from_matrix(matrix)
            total_arquetipos = len(matrix.arquetipos)

            self._view.clear_results()
            self._view.add_rows(benchmarking_rows)
            self._view.append_log(f"📊 Benchmarking generado: {total_arquetipos} arquetipos")
            self._view.set_status("Benchmarking Listo")
            self._view.set_benchmarking_state(False)
            self._view.enable_export(True)

        self._view.after(0, update_ui)

    def handle_export_benchmarking(self, folder_path: str) -> None:
        """Inicia de forma asíncrona la exportación del archivo Excel de benchmarking.

        Args:
            folder_path: Directorio de destino para guardar el reporte.
        """
        if self._current_benchmarking is None:
            self._view.after(0, lambda: self._view.set_status("Sin matriz de benchmarking"))
            return

        self._view.set_status("Exportando Benchmarking...")

        def _async_export() -> None:
            try:
                self._export_benchmarking_by_blocks(self._current_benchmarking, folder_path)
                categoria_safe = "_".join(
                    ((self._current_benchmarking.categoria or "General").strip() or "General").split()
                )
                output_file = os.path.join(folder_path, f"Benchmarking_{categoria_safe}.xlsx")
                self._view.after(0, lambda: self._view.append_log(f"🗂️ Archivo: {output_file}"))
                self._view.after(0, lambda: self._view.set_status("Benchmarking Exportado"))
            except Exception as exc:
                self._view.after(0, lambda: self._view.set_status(f"Error: {exc}"))

        self._executor.submit(_async_export)

    def _export_benchmarking_by_blocks(self, matrix: BenchmarkingMatrix, folder_path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Benchmarking"

        header_fill = PatternFill(fill_type="solid", start_color="D35400", end_color="D35400")
        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(color="2C3E50", bold=True, size=14)
        data_font = Font(color="2C3E50")

        columns = [
            "Producto (Arquetipo)",
            "Cantidad",
            "Margen Promedio",
            " ",
            "COSTO PROV.",
            "PRECIO CLI.",
            "Muestra (Casos)",
        ]

        categoria = (matrix.categoria or "").strip()
        is_macro = categoria in MACRO_CATEGORIAS
        current_row = 1

        def _write_headers(row_idx: int) -> int:
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=("" if col_name == " " else col_name))
                if col_idx == 4:
                    cell.value = ""
                    cell.fill = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
                    cell.font = Font(color="2C3E50")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    continue
                cell.fill = header_fill
                cell.font = header_font
                if col_name in {"Cantidad", "Margen Promedio", "COSTO PROV.", "PRECIO CLI.", "Muestra (Casos)"}:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            return row_idx + 1

        if is_macro:
            sub_map = {sub: [] for sub in MACRO_CATEGORIAS[categoria].get("subcategorias", {})}
            for arq in matrix.arquetipos:
                assigned = False
                for sub in sub_map.keys():
                    if self._variation_service.matches_category(f"{categoria} > {sub}", arq.nombre_arquetipo):
                        sub_map[sub].append(arq)
                        assigned = True
                        break
                if not assigned:
                    sub_map.setdefault("Otros", []).append(arq)
        else:
            sub_map = {categoria or "General": list(matrix.arquetipos)}

        for subcat, arquetipos in sub_map.items():
            if not arquetipos:
                continue

            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
            title_cell = ws.cell(row=current_row, column=1, value=subcat)
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

            current_row = _write_headers(current_row)

            i = 0
            for item in arquetipos:
                tiers = [
                    (100, item.margen_tier_100, item.costo_avg_100, item.precio_avg_100, item.casos_tier_100),
                    (500, item.margen_tier_500, item.costo_avg_500, item.precio_avg_500, item.casos_tier_500),
                    (1000, item.margen_tier_1000, item.costo_avg_1000, item.precio_avg_1000, item.casos_tier_1000),
                ]
                for cantidad, margen, costo, precio, casos in tiers:
                    fill_color = "FFCAAD" if ((i // 3) % 2) == 1 else "FFFFFF"
                    data_fill = PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color)

                    ws.cell(row=current_row, column=1, value=item.nombre_arquetipo).font = data_font

                    c2 = ws.cell(row=current_row, column=2, value=cantidad)
                    c2.font = data_font
                    c2.alignment = Alignment(horizontal="center", vertical="center")

                    c3 = ws.cell(row=current_row, column=3, value=(float(margen) / 100.0 if margen is not None else 0.0))
                    c3.number_format = "0.00%"
                    c3.font = data_font
                    c3.alignment = Alignment(horizontal="center", vertical="center")

                    ws.cell(row=current_row, column=4, value="")

                    c5 = ws.cell(row=current_row, column=5, value=round(float(costo or 0.0), 2))
                    c5.font = data_font
                    c5.alignment = Alignment(horizontal="center", vertical="center")

                    c6 = ws.cell(row=current_row, column=6, value=round(float(precio or 0.0), 2))
                    c6.font = data_font
                    c6.alignment = Alignment(horizontal="center", vertical="center")

                    c7 = ws.cell(row=current_row, column=7, value=int(casos or 0))
                    c7.font = data_font
                    c7.alignment = Alignment(horizontal="center", vertical="center")

                    for col_idx in range(1, len(columns) + 1):
                        ws.cell(row=current_row, column=col_idx).fill = data_fill

                    current_row += 1
                    i += 1

            current_row += 3

        ws.column_dimensions[get_column_letter(4)].width = 20

        for col_idx in range(1, len(columns) + 1):
            col_letter = get_column_letter(col_idx)
            if col_idx == 4:
                continue
            max_len = 0
            for cell in ws[col_letter]:
                try:
                    val = str(cell.value or "")
                except Exception:
                    val = ""
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        categoria_safe = "_".join(((matrix.categoria or "General").strip() or "General").split())
        output_file = os.path.join(folder_path, f"Benchmarking_{categoria_safe}.xlsx")
        wb.save(output_file)

    def _benchmarking_rows_from_matrix(self, matrix: BenchmarkingMatrix) -> List[ScanRow]:
        rows: List[ScanRow] = []
        for item in matrix.arquetipos:
            margen_100 = round(float(item.margen_tier_100), 2)
            margen_500 = round(float(item.margen_tier_500), 2)
            margen_1000 = round(float(item.margen_tier_1000), 2)

            rows.append(
                ScanRow(
                    fila_id=int(item.casos_tier_100),
                    articulo=item.nombre_arquetipo,
                    cantidad=100,
                    precio_prov=round(float(item.costo_avg_100), 2),
                    precio_cli=round(float(item.precio_avg_100), 2),
                    margen=margen_100,
                    arquetipo=item.nombre_arquetipo,
                    margen_fila=margen_100,
                )
            )

            rows.append(
                ScanRow(
                    fila_id=int(item.casos_tier_500),
                    articulo=item.nombre_arquetipo,
                    cantidad=500,
                    precio_prov=round(float(item.costo_avg_500), 2),
                    precio_cli=round(float(item.precio_avg_500), 2),
                    margen=margen_500,
                    arquetipo=item.nombre_arquetipo,
                    margen_fila=margen_500,
                )
            )

            rows.append(
                ScanRow(
                    fila_id=int(item.casos_tier_1000),
                    articulo=item.nombre_arquetipo,
                    cantidad=1000,
                    precio_prov=round(float(item.costo_avg_1000), 2),
                    precio_cli=round(float(item.precio_avg_1000), 2),
                    margen=margen_1000,
                    arquetipo=item.nombre_arquetipo,
                    margen_fila=margen_1000,
                )
            )
        return rows