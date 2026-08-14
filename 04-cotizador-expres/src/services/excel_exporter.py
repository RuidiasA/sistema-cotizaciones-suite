import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Servicio de exportación y maquetación comercial en formato Microsoft Excel (.xlsx).

    Transforma las estructuras de datos procesadas en propuestas comerciales formales,
    aplicando estilos corporativos, formatos numéricos monetarios y segmentación
    en dos zonas: columnas comerciales para el cliente y columnas de control para auditoría.
    """

    def __init__(self, output_dir: Optional[Union[str, Path]] = None) -> None:
        """Inicializa el exportador y asegura la existencia del directorio de salida.

        Args:
            output_dir: Ruta destino de los archivos exportados. Si es None, utiliza data/output.
        """
        if output_dir is None:
            base_module_dir = Path(__file__).resolve().parent.parent.parent
            self.output_dir = base_module_dir / "data" / "output"
        else:
            self.output_dir = Path(output_dir)

        os.makedirs(self.output_dir, exist_ok=True)

    def _generar_nombre_archivo(self, lista_pedidos: List[Dict[str, Any]]) -> str:
        """Construye un nombre de archivo normalizado y gestiona colisiones mediante sufijos de versión.

        Args:
            lista_pedidos: Resumen de pedidos que componen la cotización.

        Returns:
            Ruta absoluta o relativa del archivo Excel en formato string.
        """
        partes_nombre: List[str] = []
        for pedido in lista_pedidos:
            nombre = pedido.get("producto_nombre", "Producto")
            nombre_limpio = re.sub(r'[\\/*?:"<>|]', "", nombre).replace(" ", "_")
            partes_nombre.append(f"{pedido['cantidad']}_{nombre_limpio}")

        base_name = ", ".join(partes_nombre)
        if len(base_name) > 120:
            base_name = f"{base_name[:115]}_etc"

        filename = self.output_dir / f"{base_name}.xlsx"
        counter = 1
        while filename.exists():
            filename = self.output_dir / f"{base_name}_v{counter}.xlsx"
            counter += 1

        return str(filename)

    def exportar_cotizacion_completa(self, bloques_productos: List[Dict[str, Any]]) -> str:
        """Genera el libro de Excel con bloques independientes y diseño estructurado por producto.

        Args:
            bloques_productos: Lista de bloques de productos con sus respectivas opciones calculadas.

        Returns:
            Ruta del archivo Excel generado.
        """
        resumen_pedidos = [
            {"producto_nombre": bloque["producto_nombre"], "cantidad": bloque["cantidad"]}
            for bloque in bloques_productos
        ]
        filename = self._generar_nombre_archivo(resumen_pedidos)

        wb = Workbook()
        ws = wb.active
        ws.title = "Cotización"
        ws.views.sheetView[0].showGridLines = True

        yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        gray_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        blue_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        header_font = Font(name="Arial", size=10, bold=True, color="000000")
        title_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=10)
        foto_font = Font(name="Arial", size=9, bold=True, color="FF0000")
        control_font = Font(name="Arial", size=10, color="333333")

        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        foto_align = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

        thin_side = Side(style="thin", color="B0B0B0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        currency_fmt = '"S/." #,##0.00'

        all_headers = [
            "N°", "Proveedor", "Producto", "Foto", "Cant.",
            "Costo uni. NO IGV (S/.)", "Tiempo Entrega", "Detalle", "Costo TOTAL NO IGV (S/.)",
            "Cantidad", "Costo Prov", "Ref Real 2026"
        ]

        current_row = 2

        for bloque in bloques_productos:
            prod_nombre = bloque["producto_nombre"]
            cantidad = bloque["cantidad"]
            opciones_proveedores = bloque["opciones"]

            if not opciones_proveedores:
                continue

            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=len(all_headers)
            )
            cell_titulo = ws.cell(row=current_row, column=1, value=f"{prod_nombre.upper()} ({cantidad} UNIDADES)")
            cell_titulo.font = title_font
            cell_titulo.fill = blue_fill
            cell_titulo.alignment = center_align
            ws.row_dimensions[current_row].height = 24

            current_row += 1

            ws.row_dimensions[current_row].height = 25
            for col_idx, h_text in enumerate(all_headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=h_text)
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = yellow_fill if col_idx <= 9 else gray_fill
                cell.border = thin_border

            current_row += 1

            for opcion in opciones_proveedores:
                ws.row_dimensions[current_row].height = 110

                row_values = [
                    opcion["N°"],
                    opcion["Proveedor"],
                    opcion["Producto"],
                    opcion["Foto"],
                    opcion["Cant."],
                    opcion["Costo uni. NO IGV (S/.)"],
                    opcion["Tiempo Entrega"],
                    opcion["Detalle"],
                    opcion["Costo TOTAL NO IGV (S/.)"],
                    opcion["Cantidad_Multilinea"],
                    opcion["Costo_Prov_Multilinea"],
                    opcion["Precio_Cli_Original_2026"],
                ]

                for col_idx, val in enumerate(row_values, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.border = thin_border

                    if col_idx == 4:
                        cell.font = foto_font
                        cell.alignment = foto_align
                    elif col_idx == 8:
                        cell.font = data_font
                        cell.alignment = left_align
                    elif col_idx >= 10:
                        cell.font = control_font
                        cell.fill = gray_fill
                        cell.alignment = center_align
                    else:
                        cell.font = data_font
                        cell.alignment = center_align

                    if col_idx in [6, 9, 11, 12]:
                        cell.number_format = currency_fmt

                current_row += 1

            current_row += 2

        for col_idx in range(1, len(all_headers) + 1):
            col_letter = get_column_letter(col_idx)
            if col_idx == 8:
                ws.column_dimensions[col_letter].width = 75
            elif col_idx == 4:
                ws.column_dimensions[col_letter].width = 22
            elif col_idx in [2, 3, 6, 7, 9, 11, 12]:
                ws.column_dimensions[col_letter].width = 18
            else:
                ws.column_dimensions[col_letter].width = 12

        wb.save(filename)
        return filename