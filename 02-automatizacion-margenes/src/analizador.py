"""Módulo de Análisis Estadístico y Suavizado de Márgenes.

Aplica filtros estadísticos (IQR) por tramo de cantidad, elimina outliers
y realiza un suavizado monótono para construir curvas de margen comerciales.
Inyecta el resultado directamente en JSON y en el Excel maestro de reglas.
"""

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import numpy as np
import openpyxl
import pandas as pd

from src.limpiador import CONFIG_PRODUCTOS, clasificar_producto_estricto

ESCALAS_ESTANDAR: List[int] = [
    10, 25, 50, 100, 250, 500, 1000,
    2000, 3000, 5000, 10000, 20000, 30000, 50000, 100000
]


def ajustar_a_escala_comercial(cantidad: float) -> int:
    """Mapea una cantidad real a la escala comercial estándar más cercana.

    Args:
        cantidad: Volumen numérico detectado en el dataset.

    Returns:
        Escala comercial más próxima según el catálogo predeterminado.
    """
    return min(ESCALAS_ESTANDAR, key=lambda x: abs(x - cantidad))


def suavizar_curva_comercial(margenes_dict: Dict[str, float]) -> Dict[str, float]:
    """Aplica suavizado monótono decreciente de dos vías sobre los márgenes por volumen.

    Args:
        margenes_dict: Mapeo de escalas de volumen a márgenes porcentuales brutos.

    Returns:
        Diccionario con curvas de margen corregidas y piso mínimo de 10.0%.
    """
    cantidades = sorted([int(k) for k in margenes_dict.keys()])
    if not cantidades:
        return margenes_dict

    valores_limpios: Dict[str, float] = {str(c): margenes_dict[str(c)] for c in cantidades}

    # Corrección progresiva (Izquierda a Derecha): Previene picos al escalar volumen
    for i in range(1, len(cantidades)):
        key_prev = str(cantidades[i - 1])
        key_act = str(cantidades[i])

        if valores_limpios[key_act] > valores_limpios[key_prev]:
            ratio = cantidades[i] / cantidades[i - 1]
            factor = 1.0 - min(0.08, 0.02 * np.log2(ratio))
            valores_limpios[key_act] = round(valores_limpios[key_prev] * factor, 1)
        else:
            limite_caida_maxima = valores_limpios[key_prev] * 0.70
            if valores_limpios[key_act] < limite_caida_maxima:
                valores_limpios[key_act] = round(limite_caida_maxima, 1)

    # Corrección regresiva (Derecha a Izquierda): Amortigua valles en volúmenes bajos
    for i in range(len(cantidades) - 2, -1, -1):
        key_act = str(cantidades[i])
        key_sig = str(cantidades[i + 1])

        if valores_limpios[key_act] < valores_limpios[key_sig]:
            ratio = cantidades[i + 1] / cantidades[i]
            factor = 1.0 + min(0.08, 0.02 * np.log2(ratio))
            valores_limpios[key_act] = round(valores_limpios[key_sig] * factor, 1)

    for k in valores_limpios:
        valores_limpios[k] = max(10.0, valores_limpios[k])

    return valores_limpios


def exportar_a_excel_tarifario(matrices_finales: Dict[str, Any], ruta_excel: Path) -> None:
    """Inyecta la matriz de márgenes calculada directamente en la Hoja 1 ('Margenes_Base').

    Args:
        matrices_finales: Estructura de matrices optimizadas y suavizadas.
        ruta_excel: Ruta hacia el archivo tarifario_diseno.xlsx.
    """
    if not ruta_excel.exists():
        print(f"[WARN] No se encontró el Excel en: {ruta_excel}. Se omite la actualización de 'Margenes_Base'.")
        return

    try:
        filas_tabla: List[List[Union[str, int, float]]] = []
        for arquetipo, datos in matrices_finales.items():
            for cantidad_str, margen_val in datos.get("margenes", {}).items():
                filas_tabla.append([arquetipo, int(cantidad_str), float(margen_val)])

        wb = openpyxl.load_workbook(ruta_excel)

        if "Margenes_Base" in wb.sheetnames:
            del wb["Margenes_Base"]

        ws = wb.create_sheet(title="Margenes_Base", index=0)
        ws.append(["Producto", "Cantidad", "Margen"])

        for fila in filas_tabla:
            ws.append(fila)

        wb.save(ruta_excel)
        print(f"¡Hoja 'Margenes_Base' inyectada y actualizada en: {ruta_excel.resolve()}!")
    except Exception as exc:
        print(f"[ERROR] Ocurrió un fallo al escribir en {ruta_excel.name}: {exc}")


def analizar_y_optimizar_margenes(
    ruta_excel_entrada: Union[str, Path],
    ruta_json_salida: Optional[Union[str, Path]] = None,
    ruta_excel_tarifario: Optional[Union[str, Path]] = None
) -> Dict[str, Any]:
    """Procesa el dataset histórico, aplica filtrado IQR y compila la matriz de márgenes.

    Args:
        ruta_excel_entrada: Archivo Excel consolidado del Data Lake.
        ruta_json_salida: Ruta opcional para persistir el JSON del Módulo 02.
        ruta_excel_tarifario: Ruta opcional para inyección directa en el Módulo 03.

    Returns:
        Diccionario maestro con las matrices de margen base por arquetipo.
    """
    print("Iniciando la trituración de datos históricos con motor modular...")

    df = pd.read_excel(ruta_excel_entrada, engine="openpyxl")
    df.columns = [str(col).strip() for col in df.columns]

    col_desc = "Descripcion / Articulo"
    col_cant = "Cantidad Detectada"
    col_prov = "Costo Prov"
    col_cli = "Precio Cli"

    df[col_cant] = pd.to_numeric(df[col_cant], errors="coerce")
    df[col_prov] = pd.to_numeric(df[col_prov], errors="coerce")
    df[col_cli] = pd.to_numeric(df[col_cli], errors="coerce")
    df = df.dropna(subset=[col_cant, col_prov, col_cli])
    df = df[(df[col_prov] > 0) & (df[col_cant] > 0)]

    print("Clasificando arquetipos dinámicamente mediante el diccionario maestro...")
    df["Arquetipo_Limpio"] = [
        clasificar_producto_estricto(desc, costo)
        for desc, costo in zip(df[col_desc], df[col_prov])
    ]

    df_filtrado = df[df["Arquetipo_Limpio"].notna()].copy()

    if df_filtrado.empty:
        print("[WARN] No se encontraron registros con arquetipos válidos tras la clasificación.")
        return {}

    df_filtrado["Escala_Comercial"] = df_filtrado[col_cant].apply(ajustar_a_escala_comercial)
    df_filtrado["Margen_Calculado"] = (df_filtrado[col_cli] / df_filtrado[col_prov]) - 1
    df_filtrado = df_filtrado[df_filtrado["Margen_Calculado"] >= 0]

    matrices_crudas: Dict[str, Dict[str, float]] = {}
    grouped = df_filtrado.groupby(["Arquetipo_Limpio", "Escala_Comercial"])

    print("Agrupando bloques y ejecutando control estadístico IQR por tramo...")
    for (arquetipo, escala), tramo in grouped:
        if len(tramo) < 2:
            margen_optimo = float(tramo["Margen_Calculado"].values[0])
        else:
            q1 = tramo["Margen_Calculado"].quantile(0.25)
            q3 = tramo["Margen_Calculado"].quantile(0.75)
            iqr = q3 - q1
            tramo_filtrado = tramo[
                (tramo["Margen_Calculado"] >= (q1 - 1.5 * iqr)) &
                (tramo["Margen_Calculado"] <= (q3 + 1.5 * iqr))
            ]
            margen_optimo = (
                float(tramo_filtrado["Margen_Calculado"].median())
                if not tramo_filtrado.empty
                else float(tramo["Margen_Calculado"].median())
            )

        if arquetipo not in matrices_crudas:
            matrices_crudas[arquetipo] = {}

        matrices_crudas[arquetipo][str(int(escala))] = round(margen_optimo * 100, 1)

    print("Aplicando suavizado monótono e inyectando claves estructurales...")
    matrices_finales: Dict[str, Any] = {}

    for arquetipo, margenes_sucios in matrices_crudas.items():
        margenes_ordenados = dict(sorted(margenes_sucios.items(), key=lambda item: int(item[0])))
        margenes_limpios = suavizar_curva_comercial(margenes_ordenados)

        meta_config = CONFIG_PRODUCTOS.get(arquetipo, {
            "nombre_comercial": arquetipo,
            "macro_categoria": "MERCHANDISING",
            "subcategoria": "GENERAL",
            "filtros_prenda": [arquetipo.lower()],
            "filtros_material": []
        })

        matrices_finales[arquetipo] = {
            "nombre_comercial": meta_config["nombre_comercial"],
            "macro_categoria": meta_config.get("macro_categoria", ""),
            "subcategoria": meta_config.get("subcategoria", ""),
            "filtros_prenda": meta_config.get("filtros_prenda", []),
            "filtros_material": meta_config.get("filtros_material", []),
            "margenes": margenes_limpios
        }

    if ruta_json_salida:
        path_salida = Path(ruta_json_salida)
        path_salida.parent.mkdir(parents=True, exist_ok=True)
        with open(path_salida, "w", encoding="utf-8") as file:
            json.dump(matrices_finales, file, ensure_ascii=False, indent=2)
        print(f"¡Matriz unificada exportada con éxito en: {path_salida.resolve()}!")

    if ruta_excel_tarifario:
        exportar_a_excel_tarifario(matrices_finales, Path(ruta_excel_tarifario))

    return matrices_finales